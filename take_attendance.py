import cv2
import numpy as np
from pyzbar.pyzbar import decode
from playsound import playsound
import openpyxl
import datetime as dt

record = cv2.VideoCapture(0)
record.set(3,640)
record.set(4,480)

present_data=[]

while(1):
    s, img = record.read()
    for qr in decode(img):
        data=qr.data.decode('utf-8')
        #print(data)

        #box around
        pts = np.array([qr.polygon],np.int32)
        pts = pts.reshape((-1,1,2))
        
        #display on image
        pts2 = qr.rect
        
        if data not in present_data:
            present_data.append(data)
            #print(10*'\a')
            playsound('beep.mp3')
            color=(0,255,0)
            cv2.polylines(img,[pts],1,color,5)
            #cv2.waitKey(3)
        else:
            color=(0,0,255)    
            cv2.polylines(img,[pts],1,color,5)
            cv2.putText(img,"Thank You! {}".format(data),(pts2[0],pts2[1]),cv2.FONT_HERSHEY_PLAIN, 1,color,2)

    cv2.imshow('Result',img)
    key = cv2.waitKey(1)

    #to release camera frame
    if(key==27): #esc key value is 27
        cv2.destroyWindow("Result")
        record.release()
        break
    
print(present_data)

temp = openpyxl.load_workbook("Attendance_Sheet.xlsx")
sheet = temp.active
column = sheet.max_column
sheet.cell(1,column+1).value = dt.date.today()
for i in range(2,sheet.max_row+1):
    x=sheet.cell(i,2)
    if(x.value in present_data):
       sheet.cell(i,column+1).value="P"
    else:
        sheet.cell(i,column+1).value="Ab"

temp.save("Attendance_Sheet.xlsx")
